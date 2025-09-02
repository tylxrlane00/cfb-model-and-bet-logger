# cfb_week1_predictor_cli_v5.py
# Week-1-ready CFB score predictor with interactive prompts, situational stats, weather, CSV input, Monte Carlo, and config.

from dataclasses import dataclass
from typing import Optional, Dict, Tuple
import math
import csv
import os
import re
import json
import numpy as np
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = load_workbook = Font = None  # Excel optional

# ---------- Load Config ----------
def load_config(path="config.json"):
    """Load constants and weights from JSON config file."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Config file {path} not found. Create with constants and weights.")
    with open(path, 'r') as f:
        config = json.load(f)
    globals().update(config.get("constants", {}))
    globals()["OFF_WEIGHTS"] = config.get("weights", {}).get("off", {})
    globals()["DEF_WEIGHTS"] = config.get("weights", {}).get("def", {})
    globals()["MARKET_BLEND"] = config.get("market_blend", {})
    globals()["EV_THRESHOLDS"] = config.get("ev_thresholds", {})

# Load config at startup (assumes config.json exists)
load_config()

# ---------- CONFIG: Defaults (overridden by config.json) ----------
# League baselines (2024 FBS data: SP+, NCAA, TeamRankings)
AVG_OFF_EPA = 0.00;     STD_OFF_EPA = 0.14
AVG_DEF_EPA = 0.00;     STD_DEF_EPA = 0.14
AVG_RET_PROD = 0.65;    STD_RET_PROD = 0.11      # Adjusted from 0.68
AVG_COMPOSITE_100 = 50.0; STD_COMPOSITE_100 = 18.0
AVG_QB_100 = 50.0;      STD_QB_100 = 20.0
AVG_OL_RET = 2.8;       STD_OL_RET = 1.2
AVG_TEMPO = 65.8;       STD_TEMPO = 7.0
AVG_3D = 0.398;         STD_3D = 0.06
AVG_RZ_TD = 0.615;      STD_RZ_TD = 0.075
AVG_RZ_FG = 0.23;       STD_RZ_FG = 0.055
AVG_FD = 20.2;          STD_FD = 3.5
AVG_GIVE = 1.45;        STD_GIVE = 0.45
AVG_TAKE = 1.45;        STD_TAKE = 0.45
BASE_POINTS_PER_DRIVE = 2.28
PLAYS_PER_DRIVE = 5.8
EFF_TO_PPD_K = 0.11
SPREAD_SIGMA_PTS = 13.0
TOTALS_SIGMA_PTS = 11.0
EV_THRESHOLDS = {"strong": 0.05, "lean": 0.02}
DEFAULT_BOOK_ODDS = -110
WX_WIND_PPD_PER_MPH = -0.008
WX_WIND_TEMPO_PER_MPH = -0.0015
WX_RAIN_PPD = -0.06
WX_SNOW_PPD = -0.10
WX_RAIN_TEMPO = -0.035
WX_SNOW_TEMPO = -0.05
WX_COLD_PPD = -0.03
WX_HOT_PPD = -0.02
MARGIN_PER_NET_TO = 3.2
DEFAULT_HFA_POINTS = 2.5  # Adjusted from 2.2
MARKET_BLEND = {"use": True, "spread_weight": 0.25, "total_weight": 0.25}
OFF_WEIGHTS = {"eff": 0.45, "ret": 0.15, "transfer": 0.08, "recruit": 0.05, "qb": 0.20, "ol": 0.05, "oc_cont": 0.05}
DEF_WEIGHTS = {"eff": 0.50, "ret": 0.20, "transfer": 0.08, "recruit": 0.05, "dc_cont": 0.10, "front7": 0.10}

# ---------- Data Classes ----------
@dataclass
class TeamInput:
    name: str
    off_epa: float = 0.0
    def_epa: float = 0.0
    returning_prod_off: float = AVG_RET_PROD
    returning_prod_def: float = AVG_RET_PROD
    transfer_grade_100: float = AVG_COMPOSITE_100
    recruit_grade_100: float = AVG_COMPOSITE_100
    qb_grade_100: float = AVG_QB_100
    ol_returning_starters: int = 3
    oc_continuity: bool = True
    dc_continuity: bool = True
    front7_continuity_100: Optional[float] = None
    tempo_plays_per_game: float = AVG_TEMPO
    third_down_off: float = AVG_3D
    third_down_def: float = AVG_3D
    rz_td_off: float = AVG_RZ_TD
    rz_fg_off: float = AVG_RZ_FG
    rz_td_def: float = AVG_RZ_TD
    rz_fg_def: float = AVG_RZ_FG
    first_downs_pg: float = AVG_FD
    giveaways_pg: float = AVG_GIVE
    takeaways_pg: float = AVG_TAKE
    sos_factor: float = 0.0  # New: Strength of Schedule adjustment

@dataclass
class GameWeather:
    is_indoor: bool = False
    temp_f: float = 70.0
    wind_mph: float = 5.0
    precip: str = "none"

# ---------- Core Math ----------
def z(v: float, m: float, s: float) -> float:
    s = s if s and s > 1e-9 else 1.0
    return (v - m) / s

def preseason_off_index(t: TeamInput) -> float:
    eff_z = z(t.off_epa + t.sos_factor * STD_OFF_EPA, AVG_OFF_EPA, STD_OFF_EPA)
    ret_z = z(t.returning_prod_off, AVG_RET_PROD, STD_RET_PROD)
    tr_z = z(t.transfer_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    rec_z = z(t.recruit_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    qb_z = z(t.qb_grade_100, AVG_QB_100, STD_QB_100)
    ol_z = z(float(t.ol_returning_starters), AVG_OL_RET, STD_OL_RET)
    oc = 1.0 if t.oc_continuity else 0.0
    return (OFF_WEIGHTS["eff"]*eff_z + OFF_WEIGHTS["ret"]*ret_z + OFF_WEIGHTS["transfer"]*tr_z +
            OFF_WEIGHTS["recruit"]*rec_z + OFF_WEIGHTS["qb"]*qb_z + OFF_WEIGHTS["ol"]*ol_z +
            OFF_WEIGHTS["oc_cont"]*oc)

def preseason_def_index(t: TeamInput) -> float:
    eff_component = -(t.def_epa - t.sos_factor * STD_DEF_EPA)  # SOS adjusts def_epa
    eff_z = z(eff_component, -AVG_DEF_EPA, STD_DEF_EPA)
    ret_z = z(t.returning_prod_def, AVG_RET_PROD, STD_RET_PROD)
    tr_z = z(t.transfer_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    rec_z = z(t.recruit_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    dc = 1.0 if t.dc_continuity else 0.0
    f7_z = 0.0 if t.front7_continuity_100 is None else z(t.front7_continuity_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    return (DEF_WEIGHTS["eff"]*eff_z + DEF_WEIGHTS["ret"]*ret_z + DEF_WEIGHTS["transfer"]*tr_z +
            DEF_WEIGHTS["recruit"]*rec_z + DEF_WEIGHTS["dc_cont"]*dc + DEF_WEIGHTS["front7"]*f7_z)

def combined_plays(base_tempo_home: float, base_tempo_away: float, home_fd: float, away_fd: float, wx: GameWeather) -> float:
    raw = 0.96 * (base_tempo_home + base_tempo_away)
    fd_bump = 1.0 + 0.004 * ((home_fd - AVG_FD) + (away_fd - AVG_FD))
    raw *= max(0.85, min(1.15, fd_bump))
    if not wx.is_indoor:
        wind_factor = 1.0 + WX_WIND_TEMPO_PER_MPH * max(0.0, wx.wind_mph)
        raw *= max(0.85, wind_factor)
        if wx.precip in ("rain", "snow"):
            raw *= (1.0 + (WX_RAIN_TEMPO if wx.precip == "rain" else WX_SNOW_TEMPO))
    return raw

def weather_ppd_multiplier(wx: GameWeather) -> float:
    if wx.is_indoor:
        return 1.0
    mult = 1.0
    wind_pen = WX_WIND_PPD_PER_MPH * max(0.0, wx.wind_mph)
    mult *= max(0.70, 1.0 + wind_pen)
    if wx.precip == "rain":
        mult *= (1.0 + WX_RAIN_PPD)
    elif wx.precip == "snow":
        mult *= (1.0 + WX_SNOW_PPD)
    if wx.temp_f < 40:
        mulat *= (1.0 + WX_COLD_PPD)
    elif wx.temp_f > 90:
        mult *= (1.0 + WX_HOT_PPD)
    return mult

def situational_ppd_multiplier(off: TeamInput, opp_def: TeamInput) -> float:
    def z(v, m, s): s = s if s and s > 1e-9 else 1.0; return (v - m) / s
    m_3d = 1.0 + 0.05 * (z(off.third_down_off, AVG_3D, STD_3D) - z(opp_def.third_down_def, AVG_3D, STD_3D))
    m_3d = max(0.85, min(1.15, m_3d))
    e_off = 7.0 * max(0, min(1, off.rz_td_off)) + 3.0 * max(0, min(1, off.rz_fg_off))
    e_def_allowed = 7.0 * max(0, min(1, opp_def.rz_td_def)) + 3.0 * max(0, min(1, opp_def.rz_fg_def))
    rz_factor = (e_off / 4.8) * (4.8 / max(1e-6, e_def_allowed)) ** 0.5  # Fixed: normalize vs avg
    m_rz = max(0.85, min(1.15, rz_factor))
    give_pen = max(0.0, z(off.giveaways_pg, AVG_GIVE, STD_GIVE))
    take_pen = max(0.0, z(opp_def.takeaways_pg, AVG_TAKE, STD_TAKE))
    m_to = max(0.90, 1.0 - 0.02 * give_pen - 0.015 * take_pen)
    return m_3d * m_rz * m_to

def expected_net_turnovers(home: TeamInput, away: TeamInput) -> float:
    return (home.takeaways_pg + away.giveaways_pg) - (home.giveaways_pg + away.takeaways_pg)

def model_points_per_drive(off_idx: float, opp_def_idx: float) -> float:
    gap = off_idx - opp_def_idx
    return BASE_POINTS_PER_DRIVE * math.exp(EFF_TO_PPD_K * gap)

def predict_game(
    home: TeamInput,
    away: TeamInput,
    weather: GameWeather,
    use_preseason: bool = True,
    hfa_points: float = DEFAULT_HFA_POINTS,
    vegas_spread: Optional[float] = None,
    vegas_total: Optional[float] = None,
    use_market_blend: bool = MARKET_BLEND["use"],
    spread_weight: float = MARKET_BLEND["spread_weight"],
    total_weight: float = MARKET_BLEND["total_weight"],
) -> Dict[str, float]:
    if use_preseason:
        home_off_idx = preseason_off_index(home)
        home_def_idx = preseason_def_index(home)
        away_off_idx = preseason_off_index(away)
        away_def_idx = preseason_def_index(away)
    else:
        home_off_idx = z(home.off_epa + home.sos_factor * STD_OFF_EPA, AVG_OFF_EPA, STD_OFF_EPA)
        away_off_idx = z(away.off_epa + away.sos_factor * STD_OFF_EPA, AVG_OFF_EPA, STD_OFF_EPA)
        home_def_idx = z(-(home.def_epa - home.sos_factor * STD_DEF_EPA), -AVG_DEF_EPA, STD_DEF_EPA)
        away_def_idx = z(-(away.def_epa - away.sos_factor * STD_DEF_EPA), -AVG_DEF_EPA, STD_DEF_EPA)

    min_tempo = 50.0
    plays = combined_plays(max(min_tempo, home.tempo_plays_per_game), max(min_tempo, away.tempo_plays_per_game),
                           home.first_downs_pg, away.first_downs_pg, weather)
    total_drives = max(10.0, plays / PLAYS_PER_DRIVE)
    den = max(1e-6, home.tempo_plays_per_game + away.tempo_plays_per_game)
    drives_home = total_drives * (home.tempo_plays_per_game / den)
    drives_away = total_drives - drives_home

    home_ppd = model_points_per_drive(home_off_idx, away_def_idx)
    away_ppd = model_points_per_drive(away_off_idx, home_def_idx)

    wx_mult = weather_ppd_multiplier(weather)
    home_mult = situational_ppd_multiplier(home, away) * wx_mult
    away_mult = situational_ppd_multiplier(away, home) * wx_mult
    home_ppd *= home_mult
    away_ppd *= away_mult

    home_pts = home_ppd * drives_home
    away_pts = away_ppd * drives_away

    model_total = home_pts + away_pts
    model_spread = (home_pts - away_pts) + hfa_points

    net_to = expected_net_turnovers(home, away)
    model_spread += MARGIN_PER_NET_TO * net_to

    home_pts_adj = (model_total + model_spread) / 2.0
    away_pts_adj = model_total - home_pts_adj

    final_spread = model_spread
    final_total = model_total
    if use_market_blend and (vegas_spread is not None or vegas_total is not None):
        if vegas_spread is not None:
            final_spread = (1 - spread_weight) * model_spread + spread_weight * vegas_spread
        if vegas_total is not None:
            final_total = (1 - total_weight) * model_total + total_weight * vegas_total
        home_pts_adj = (final_total + final_spread) / 2.0
        away_pts_adj = final_total - home_pts_adj

    home_pts_adj = max(0.0, home_pts_adj)
    away_pts_adj = max(0.0, away_pts_adj)

    return {
        "home_team": home.name,
        "away_team": away.name,
        "home_points": round(home_pts_adj, 1),
        "away_points": round(away_pts_adj, 1),
        "spread_home_minus_away": round(final_spread, 1),
        "total_points": round(final_total, 1),
        "model_spread_no_market": round(model_spread, 1),
        "model_total_no_market": round(model_total, 1),
        "vegas_spread": vegas_spread if vegas_spread is not None else "",
        "vegas_total": vegas_total if vegas_total is not None else "",
        "hfa_points": hfa_points,
        "used_preseason": use_preseason,
        "combined_plays_est": round(plays, 1),
        "home_ppd": round(home_ppd, 3),
        "away_ppd": round(away_ppd, 3),
        "drives_home": round(drives_home, 1),
        "drives_away": round(drives_away, 1),
        "net_turnovers_home_adv": round(net_to, 2),
        "wx_indoor": weather.is_indoor,
        "wx_temp_f": weather.temp_f,
        "wx_wind_mph": weather.wind_mph,
        "wx_precip": weather.precip,
    }

# ---------- Monte Carlo ----------
def monte_carlo(pred: Dict, n_sims=1000) -> Dict:
    """Run Monte Carlo sims for win prob and score distributions."""
    home_scores, away_scores = [], []
    for _ in range(n_sims):
        home_ppd_var = np.random.normal(pred['home_ppd'], STD_OFF_EPA * EFF_TO_PPD_K)
        away_ppd_var = np.random.normal(pred['away_ppd'], STD_OFF_EPA * EFF_TO_PPD_K)
        net_to_var = np.random.normal(pred['net_turnovers_home_adv'], STD_GIVE)
        home_pts = max(0, home_ppd_var * pred['drives_home'] + MARGIN_PER_NET_TO * net_to_var + pred['hfa_points'])
        away_pts = max(0, away_ppd_var * pred['drives_away'])
        home_scores.append(home_pts)
        away_scores.append(away_pts)
    home_win_prob = sum(h > a for h, a in zip(home_scores, away_scores)) / n_sims
    return {
        "home_win_prob": home_win_prob,
        "sim_home_pts": np.mean(home_scores),
        "sim_away_pts": np.mean(away_scores),
        "sim_home_pts_50th": np.percentile(home_scores, 50),
        "sim_away_pts_50th": np.percentile(away_scores, 50),
    }

# ---------- CSV Input ----------
def load_team_from_csv(filename: str, team_name: str) -> TeamInput:
    """Load team stats from CSV."""
    try:
        with open(filename, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['Team'].lower() == team_name.lower():
                    return TeamInput(
                        name=row['Team'],
                        off_epa=float(row.get('Off EPA/play', 0.0)),
                        def_epa=float(row.get('Def EPA/play', 0.0)),
                        returning_prod_off=float(row.get('Returning Production Offense (0–1)', AVG_RET_PROD)),
                        returning_prod_def=float(row.get('Returning Production Defense (0–1)', AVG_RET_PROD)),
                        transfer_grade_100=float(row.get('Transfer Portal Impact (0–100)', AVG_COMPOSITE_100)),
                        recruit_grade_100=float(row.get('Recruiting / Roster Quality Grade (0–100)', AVG_COMPOSITE_100)),
                        qb_grade_100=float(row.get('QB Grade (0–100)', AVG_QB_100)),
                        ol_returning_starters=int(row.get('Returning OL Starters', 3)),
                        oc_continuity=row.get('OC Continuity', 'True').lower() == 'true',
                        dc_continuity=row.get('DC Continuity', 'True').lower() == 'true',
                        front7_continuity_100=float(row.get('Front-7 Continuity Grade (0–100)')) if row.get('Front-7 Continuity Grade (0–100)') else None,
                        tempo_plays_per_game=float(row.get('Offensive Tempo (plays/game)', AVG_TEMPO)),
                        third_down_off=float(row.get('3rd Down Conv % (Offense)', AVG_3D)),
                        third_down_def=float(row.get('3rd Down Conv % (Defense)', AVG_3D)),
                        rz_td_off=float(row.get('Red Zone Offense – TD %', AVG_RZ_TD)),
                        rz_fg_off=float(row.get('Red Zone Offense – FG %', AVG_RZ_FG)),
                        rz_td_def=float(row.get('Red Zone Defense – TD %', AVG_RZ_TD)),
                        rz_fg_def=float(row.get('Red Zone Defense – FG %', AVG_RZ_FG)),
                        first_downs_pg=float(row.get('First Downs per Game (Offense)', AVG_FD)),
                        giveaways_pg=float(row.get('Giveaways per Game', AVG_GIVE)),
                        takeaways_pg=float(row.get('Takeaways per Game', AVG_TAKE)),
                        sos_factor=float(row.get('sos_factor', 0.0))
                    )
            raise ValueError(f"Team '{team_name}' not found in {filename}. Available teams: {[row['Team'] for row in csv.DictReader(open(filename, 'r'))]}")
    except FileNotFoundError:
        raise FileNotFoundError(f"CSV file {filename} not found.")
    except Exception as e:
        raise ValueError(f"Error reading {filename}: {e}")

# ---------- CLI Helpers ----------
def _ask(prompt: str, cast, default, min_val=None, max_val=None):
    """Ask for input with range validation."""
    while True:
        raw = input(f"{prompt} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            val = cast(raw)
            if min_val is not None and val < min_val:
                print(f"  ✖ Value must be >= {min_val}. Try again.")
                continue
            if max_val is not None and val > max_val:
                print(f"  ✖ Value must be <= {max_val}. Try again.")
                continue
            return val
        except ValueError:
            print("  ✖ Invalid input. Try again.")

def _ask_bool(prompt: str, default: bool) -> bool:
    dstr = "Y" if default else "N"
    while True:
        raw = input(f"{prompt} (Y/N) [{dstr}]: ").strip().lower()
        if raw == "":
            return default
        if raw in ("y", "yes"): return True
        if raw in ("n", "no"): return False
        print("  ✖ Please enter Y or N.")

def _ask_optional_float(prompt: str, min_val=None, max_val=None) -> Optional[float]:
    while True:
        raw = input(f"{prompt} (or Enter to skip): ").strip()
        if raw == "":
            return None
        try:
            val = float(raw)
            if min_val is not None and val < min_val:
                print(f"  ✖ Value must be >= {min_val}. Try again.")
                continue
            if max_val is not None and val > max_val:
                print(f"  ✖ Value must be <= {max_val}. Try again.")
                continue
            return val
        except ValueError:
            print("  ✖ Invalid number. Try again or press Enter to skip.")

def _rate(prompt: str, default: float) -> float:
    return _ask(prompt, float, default, min_val=0.0, max_val=1.0)

def _ask_spread(home_team_name: str, away_team_name: str):
    def explain(val: float):
        if abs(val) < 1e-9:
            print("  → Interpreted as Home–Away = +0 (Pick’em)")
            print("    Explanation: Pick’em")
        elif val > 0:
            print(f"  → Interpreted as Home–Away = +{val:g}")
            print(f"    Explanation: {home_team_name} favored by {val:g}")
        else:
            print(f"  → Interpreted as Home–Away = {val:g}")
            print(f"    Explanation: {away_team_name} favored by {abs(val):g}")
    
    def has_word(text: str, word: str) -> bool:
        word = word.strip().lower()
        text = text.lower()
        if " " in word:
            return word in text
        return re.search(r'\b' + re.escape(word) + r'\b', text) is not None

    while True:
        raw = input(
            f"Point spread (e.g., '{home_team_name} -7.5', '{away_team_name} +3', 'pk', or raw +/- like +7.5) [Enter to skip]: "
        ).strip()
        if raw == "":
            return None
        rl = raw.lower()
        if rl in {"pk", "pick", "pickem", "pick'em", "0", "0.0"}:
            val = 0.0
            explain(val)
            return val
        m = re.search(r'([+-]?\d+(?:\.\d+)?)', rl)
        if not m:
            print("  ✖ Couldn't find a number. Try again.")
            continue
        num = float(m.group(1))
        is_home = has_word(rl, "home") or has_word(rl, home_team_name)
        is_away = has_word(rl, "away") or has_word(rl, away_team_name)
        if is_home and is_away:
            print("  ✖ Please reference only one side (Home or Away or one team name), not both.")
            continue
        if not (is_home or is_away):
            val = num
            explain(val)
            return val
        if is_home:
            val = +abs(num) if num < 0 else -abs(num)
            explain(val)
            return val
        if is_away:
            val = -abs(num) if num < 0 else +abs(num)
            explain(val)
            return val

def _ask_total():
    while True:
        raw = input("Game total (e.g., 'o/u 54.5' or '54.5') [Enter to skip]: ").strip().lower()
        if raw == "":
            return None
        m = re.search(r'(\d+(?:\.\d+)?)', raw)
        if not m:
            print("  ✖ Couldn't find a number. Try again or press Enter to skip.")
            continue
        total = float(m.group(1))
        if total <= 0:
            print("  ✖ Total must be positive.")
            continue
        print(f"  → Interpreted total = {total:g}")
        return total

def _ask_team(which: str) -> TeamInput:
    print(f"\n--- Enter {which} TEAM info ---")
    name = input("Team name: ").strip() or which
    off_epa = _ask("Last season OFF EPA/play", float, 0.0, min_val=-1.0, max_val=1.0)
    def_epa = _ask("Last season DEF EPA/play (negative is better)", float, 0.0, min_val=-1.0, max_val=1.0)
    returning_prod_off = _rate("Returning production OFF (0-1)", AVG_RET_PROD)
    returning_prod_def = _rate("Returning production DEF (0-1)", AVG_RET_PROD)
    transfer_grade_100 = _ask("Transfer/portal impact grade (0-100)", float, AVG_COMPOSITE_100, min_val=0, max_val=100)
    recruit_grade_100 = _ask("Recruiting/roster quality grade (0-100)", float, AVG_COMPOSITE_100, min_val=0, max_val=100)
    qb_grade_100 = _ask("QB grade (0-100)", float, AVG_QB_100, min_val=0, max_val=100)
    ol_returning_starters = _ask("Returning OL starters (0-5)", int, 3, min_val=0, max_val=5)
    oc_continuity = _ask_bool("OC continuity from last year?", True)
    dc_continuity = _ask_bool("DC continuity from last year?", True)
    front7 = _ask_optional_float("Front-7 continuity grade (0-100)", min_val=0, max_val=100)
    tempo = _ask("Offensive tempo (plays/game, ~60-75)", float, AVG_TEMPO, min_val=50, max_val=80)
    third_off = _rate("Offense 3rd-down rate (0-1 or %)", AVG_3D)
    third_def = _rate("Defense 3rd-down allowed (0-1 or %)", AVG_3D)
    rz_td_off = _rate("Offense Red-Zone TD rate (0-1 or %)", AVG_RZ_TD)
    rz_fg_off = _rate("Offense Red-Zone FG rate (0-1 or %)", AVG_RZ_FG)
    rz_td_def = _rate("Defense RZ TD allowed (0-1 or %)", AVG_RZ_TD)
    rz_fg_def = _rate("Defense RZ FG allowed (0-1 or %)", AVG_RZ_FG)
    first_downs = _ask("First downs per game", float, AVG_FD, min_val=10, max_val=30)
    giveaways = _ask("Giveaways per game", float, AVG_GIVE, min_val=0, max_val=5)
    takeaways = _ask("Takeaways per game", float, AVG_TAKE, min_val=0, max_val=5)
    sos_factor = _ask("SOS adjustment factor (-1 to 1, e.g., 0.5 for tough schedule)", float, 0.0, min_val=-1.0, max_val=1.0)

    return TeamInput(
        name=name, off_epa=off_epa, def_epa=def_epa,
        returning_prod_off=returning_prod_off, returning_prod_def=returning_prod_def,
        transfer_grade_100=transfer_grade_100, recruit_grade_100=recruit_grade_100,
        qb_grade_100=qb_grade_100, ol_returning_starters=ol_returning_starters,
        oc_continuity=oc_continuity, dc_continuity=dc_continuity,
        front7_continuity_100=front7, tempo_plays_per_game=tempo,
        third_down_off=third_off, third_down_def=third_def,
        rz_td_off=rz_td_off, rz_fg_off=rz_fg_off, rz_td_def=rz_td_def, rz_fg_def=rz_fg_def,
        first_downs_pg=first_downs, giveaways_pg=giveaways, takeaways_pg=takeaways,
        sos_factor=sos_factor
    )

def _print_team_summary(t: TeamInput, which: str):
    print(f"\n[{which} SUMMARY] {t.name}")
    print(f"  EPA: OFF {t.off_epa} | DEF {t.def_epa} | SOS Factor {t.sos_factor}")
    print(f"  Returning: OFF {t.returning_prod_off} | DEF {t.returning_prod_def}")
    print(f"  Transfer {t.transfer_grade_100} | Recruit {t.recruit_grade_100} | QB {t.qb_grade_100} | OL {t.ol_returning_starters}")
    print(f"  OC cont {t.oc_continuity} | DC cont {t.dc_continuity} | Front7 {t.front7_continuity_100}")
    print(f"  Tempo {t.tempo_plays_per_game}")
    print(f"  3rd: OFF {t.third_down_off} | DEF {t.third_down_def}")
    print(f"  RZ OFF: TD {t.rz_td_off} FG {t.rz_fg_off} | RZ DEF: TD {t.rz_td_def} FG {t.rz_fg_def}")
    print(f"  First downs {t.first_downs_pg} | Giveaways {t.giveaways_pg} | Takeaways {t.takeaways_pg}")

def _print_wx_summary(wx: GameWeather):
    print("\n[WEATHER SUMMARY]")
    print(f"  Indoor {wx.is_indoor} | Temp {wx.temp_f}F | Wind {wx.wind_mph} mph | Precip {wx.precip}")

def _ask_weather() -> GameWeather:
    print("\n--- Weather / Venue ---")
    indoor = _ask_bool("Is the game indoors/under a roof?", False)
    temp = _ask("Temp (°F)", float, 70.0, min_val=0, max_val=120)
    wind = _ask("Wind (mph)", float, 5.0, min_val=0, max_val=50)
    precip = input("Precip (none/light/rain/snow) [none]: ").strip().lower() or "none"
    if precip not in ("none", "light", "rain", "snow"):
        precip = "none"
    return GameWeather(is_indoor=indoor, temp_f=temp, wind_mph=wind, precip=precip)

def maybe_save_outputs(pred: dict):
    folder = r"C:\Users\User\OneDrive\Desktop\CFP Betting Model\Scripts\PredictionResults"
    os.makedirs(folder, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    csv_path = os.path.join(folder, f"cfb_predictions_{date_str}.csv")
    xlsx_path = os.path.join(folder, f"cfb_predictions_{date_str}.xlsx")
    preferred = [
        "away_team", "away_points",
        "home_team", "home_points",
        "spread_home_minus_away", "total_points",
        "vegas_spread", "vegas_total",
        "model_spread_no_market", "model_total_no_market",
        "hfa_points", "used_preseason",
        "net_turnovers_home_adv",
        "combined_plays_est", "home_ppd", "away_ppd",
        "drives_home", "drives_away",
        "wx_indoor", "wx_temp_f", "wx_wind_mph", "wx_precip",
        "home_win_prob", "sim_home_pts", "sim_away_pts",
        "sig_spread_line", "sig_spread_model", "sig_spread_edge_pts", "sig_spread_p_home_covers", "sig_spread_ev_pct", "sig_spread_side", "sig_spread_reco",
        "sig_total_line", "sig_total_model", "sig_total_edge_pts", "sig_total_p_over", "sig_total_ev_pct", "sig_total_side", "sig_total_reco",
        "sig_ml_home_model_p", "sig_ml_home_break_even", "sig_ml_home_ev_pct", "sig_ml_home_signal",
        "sig_ml_away_model_p", "sig_ml_away_break_even", "sig_ml_away_ev_pct", "sig_ml_away_signal"
    ]
    pred_keys = list(pred.keys())
    header = [k for k in preferred if k in pred_keys]
    header += [k for k in pred_keys if k not in header]
    def fmt(x):
        if isinstance(x, float):
            return f"{x:.2f}"
        return x
    is_new_csv = not os.path.exists(csv_path)
    with open(csv_path, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        if is_new_csv:
            writer.writeheader()
        writer.writerow({k: fmt(pred.get(k, "")) for k in header})
    print(f"  ✓ Saved CSV to {csv_path}")
    if Workbook is None or load_workbook is None:
        return
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    existing_header = [c.value for c in ws[1]]
    if existing_header != header:
        union = existing_header[:] if existing_header else []
        for k in header:
            if k not in union:
                union.append(k)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.remove(ws)
        ws = wb.create_sheet("Predictions", 0)
        ws.append(union)
        for r in rows:
            row_dict = dict(zip(existing_header, r))
            ws.append([fmt(row_dict.get(k, "")) for k in union])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        header = union
    ws.append([fmt(pred.get(k, "")) for k in header])
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)
    wb.save(xlsx_path)
    print(f"  ✓ Saved pretty Excel to {xlsx_path}")

def _phi(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))

def american_to_decimal(odds: float) -> float:
    if odds >= 100:
        return 1.0 + odds / 100.0
    elif odds <= -100:
        return 1.0 + 100.0 / abs(odds)
    else:
        raise ValueError("American odds should be ≤ -100 or ≥ +100")

def break_even_prob(odds: float) -> float:
    dec = american_to_decimal(odds)
    return 1.0 / dec

def expected_value_pct(p: float, odds: float) -> float:
    dec = american_to_decimal(odds)
    return p * (dec - 1.0) - (1.0 - p)

def tier_from_ev(ev_pct: float) -> str:
    if ev_pct >= EV_THRESHOLDS["strong"]:
        return "STRONG"
    if ev_pct >= EV_THRESHOLDS["lean"]:
        return "LEAN"
    return "PASS"

def spread_signal(model_spread: float, line_home_minus_away: float, odds: float = DEFAULT_BOOK_ODDS):
    edge_pts = model_spread - line_home_minus_away
    z = edge_pts / SPREAD_SIGMA_PTS
    p_home_covers = _phi(z)
    be = break_even_prob(odds)
    ev = expected_value_pct(p_home_covers, odds)
    side = "HOME" if p_home_covers > be else "AWAY"
    signal = tier_from_ev(abs(ev)) if ev > 0 else "PASS"
    return {
        "market_line": line_home_minus_away,
        "model_spread": round(model_spread, 2),
        "edge_pts": round(edge_pts, 2),
        "p_home_covers": round(p_home_covers, 3),
        "odds": odds,
        "break_even": round(be, 3),
        "ev_pct": round(ev, 3),
        "side": side,
        "signal": signal
    }

def total_signal(model_total: float, line_total: float, odds: float = DEFAULT_BOOK_ODDS):
    edge_pts = model_total - line_total
    z = edge_pts / TOTALS_SIGMA_PTS
    p_over = _phi(z)
    be = break_even_prob(odds)
    ev = expected_value_pct(p_over, odds)
    side = "OVER" if p_over > be else "UNDER"
    signal = tier_from_ev(abs(ev)) if ev > 0 else "PASS"
    return {
        "market_total": line_total,
        "model_total": round(model_total, 2),
        "edge_pts": round(edge_pts, 2),
        "p_over": round(p_over, 3),
        "odds": odds,
        "break_even": round(be, 3),
        "ev_pct": round(ev, 3),
        "side": side,
        "signal": signal
    }

def moneyline_signal(model_home_win_prob: float, odds_home: float = None, odds_away: float = None):
    out = {}
    if odds_home is not None:
        be_h = break_even_prob(odds_home)
        ev_h = expected_value_pct(model_home_win_prob, odds_home)
        out["HOME"] = {
            "model_p": round(model_home_win_prob, 3),
            "odds": odds_home,
            "break_even": round(be_h, 3),
            "ev_pct": round(ev_h, 3),
            "signal": tier_from_ev(ev_h) if ev_h > 0 else "PASS",
        }
    if odds_away is not None:
        p_away = 1.0 - model_home_win_prob
        be_a = break_even_prob(odds_away)
        ev_a = expected_value_pct(p_away, odds_away)
        out["AWAY"] = {
            "model_p": round(p_away, 3),
            "odds": odds_away,
            "break_even": round(be_a, 3),
            "ev_pct": round(ev_a, 3),
            "signal": tier_from_ev(ev_a) if ev_a > 0 else "PASS",
        }
    return out

def run_once():
    try:
        team_csv = input("Enter team stats CSV file path (or press Enter for manual entry): ").strip()
        if team_csv:
            try:
                with open(team_csv, 'r') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # Read header row
                    teams = [row[0] for row in reader]
                    print("Available teams:", teams)
            except FileNotFoundError:
                print(f"Error: CSV file '{team_csv}' not found. Please check the file path.")
                return
            except Exception as e:
                print(f"Error reading CSV file: {e}")
                return
            home_name = input("Home team name: ").strip()
            away_name = input("Away team name: ").strip()
            try:
                home = load_team_from_csv(team_csv, home_name)
                away = load_team_from_csv(team_csv, away_name)
            except ValueError as e:
                print(f"Error: {e}")
                return
            except Exception as e:
                print(f"Error loading teams: {e}")
                return
        else:
            home = _ask_team("HOME")
            away = _ask_team("AWAY")
        wx = _ask_weather()

        while True:
            print("\n=== Review your inputs ===")
            _print_team_summary(home, "HOME")
            _print_team_summary(away, "AWAY")
            _print_wx_summary(wx)
            choice = input("\nLooks good? Type 'ok' to continue, or type 'home', 'away', or 'wx' to re-enter that section (or 'restart' to start over): ").strip().lower()
            if choice in ("ok", ""):
                break
            elif choice == "home":
                home = _ask_team("HOME") if not team_csv else load_team_from_csv(team_csv, input("Home team name: ").strip())
            elif choice == "away":
                away = _ask_team("AWAY") if not team_csv else load_team_from_csv(team_csv, input("Away team name: ").strip())
            elif choice in ("wx", "weather"):
                wx = _ask_weather()
            elif choice == "restart":
                team_csv = input("Enter team stats CSV file path (or press Enter for manual entry): ").strip()
                if team_csv:
                    try:
                        with open(team_csv, 'r') as f:
                            reader = csv.reader(f)
                            headers = next(reader)
                            teams = [row[0] for row in reader]
                            print("Available teams:", teams)
                    except FileNotFoundError:
                        print(f"Error: CSV file '{team_csv}' not found. Please check the file path.")
                        return
                    except Exception as e:
                        print(f"Error reading CSV file: {e}")
                        return
                    home_name = input("Home team name: ").strip()
                    away_name = input("Away team name: ").strip()
                    try:
                        home = load_team_from_csv(team_csv, home_name)
                        away = load_team_from_csv(team_csv, away_name)
                    except ValueError as e:
                        print(f"Error: {e}")
                        return
                    except Exception as e:
                        print(f"Error loading teams: {e}")
                        return
                else:
                    home = _ask_team("HOME")
                    away = _ask_team("AWAY")
                wx = _ask_weather()
            else:
                print("  ✖ Not a valid option. Type ok/home/away/wx/restart.")

        print("\n--- Game context ---")
        hfa = _ask("Home-field advantage in points", float, DEFAULT_HFA_POINTS, min_val=0, max_val=10)
        use_preseason = _ask_bool("Use preseason components (EPA/ret/recruit/etc.)?", True)
        use_market = _ask_bool("Blend with Vegas numbers if provided?", MARKET_BLEND["use"])

        vegas_spread = None
        vegas_total = None
        odds_home = None
        odds_away = None
        spread_w = MARKET_BLEND["spread_weight"]
        total_w = MARKET_BLEND["total_weight"]

        if use_market:
            print("\nSpread entry is flexible. Examples:")
            print(f"  • '{home.name} -7.5'  (home favored by 7.5)")
            print(f"  • '{away.name} +3'    (away is +3 underdog → home favored by 3)")
            print("  • 'pk' or '0'         (pick'em)")
            print("  • Or raw Home–Away number: +7.5 or -3.5")
            vegas_spread = _ask_spread(home.name, away.name)
            vegas_total = _ask_total()
            spread_w = _ask("Blend weight toward Vegas spread (0-1)", float, spread_w, min_val=0, max_val=1)
            total_w = _ask("Blend weight toward Vegas total (0-1)", float, total_w, min_val=0, max_val=1)
            odds_home = _ask_optional_float("Home ML odds (e.g., -150)", min_val=-10000, max_val=10000)
            odds_away = _ask_optional_float("Away ML odds (e.g., +200)", min_val=-10000, max_val=10000)

        pred = predict_game(
            home, away, wx,
            use_preseason=use_preseason,
            hfa_points=hfa,
            vegas_spread=vegas_spread, vegas_total=vegas_total,
            use_market_blend=use_market, spread_weight=spread_w, total_weight=total_w
        )

        sim = monte_carlo(pred)
        pred.update({
            "home_win_prob": round(sim["home_win_prob"], 3),
            "sim_home_pts": round(sim["sim_home_pts"], 1),
            "sim_away_pts": round(sim["sim_away_pts"], 1),
            "sim_home_pts_50th": round(sim["sim_home_pts_50th"], 1),
            "sim_away_pts_50th": round(sim["sim_away_pts_50th"], 1)
        })

        print("\n=== Prediction ===")
        print(f"Final projected score: {pred['away_team']} {pred['away_points']} @ {pred['home_team']} {pred['home_points']}")
        print(f"Home win probability: {pred['home_win_prob']*100:.1f}%")
        print(f"Simulated avg: Home {pred['sim_home_pts']} | Away {pred['sim_away_pts']}")
        print(f"Simulated median: Home {pred['sim_home_pts_50th']} | Away {pred['sim_away_pts_50th']}")

        print("\n--- Vegas-adjusted (if blended) ---")
        print(f"Spread (Home − Away): {pred['spread_home_minus_away']}   |   Total: {pred['total_points']}")

        print("\n--- Model only (no Vegas blend) ---")
        print(f"Spread (Home − Away): {pred['model_spread_no_market']}   |   Total: {pred['model_total_no_market']}")

        print("\n--- Diagnostics ---")
        print(f"Combined plays est: {pred['combined_plays_est']}")
        print(f"Home PPD: {pred['home_ppd']}   |   Away PPD: {pred['away_ppd']}")
        print(f"Drives: home {pred['drives_home']}  |  away {pred['drives_away']}")
        print(f"Expected net TO (home advantaged): {pred['net_turnovers_home_adv']}")
        print(f"WX: indoor={pred['wx_indoor']} temp={pred['wx_temp_f']}F wind={pred['wx_wind_mph']}mph precip={pred['wx_precip']}")

        if vegas_spread is not None:
            sig_spread = spread_signal(pred["model_spread_no_market"], vegas_spread, DEFAULT_BOOK_ODDS)
            vegas_fav = home.name if vegas_spread > 0 else away.name
            vegas_line_txt = f"{vegas_fav} {abs(vegas_spread):.1f}"
            model_fav = home.name if sig_spread["model_spread"] > 0 else away.name
            model_line_txt = f"{model_fav} {abs(sig_spread['model_spread']):.1f}"
            rec_team = home.name if sig_spread["side"] == "HOME" else away.name
            p_rec_covers = sig_spread["p_home_covers"] if sig_spread["side"] == "HOME" else (1.0 - sig_spread["p_home_covers"])
            print("\n--- Spread Signal ---")
            print(f"Vegas line: {vegas_line_txt}")
            print(f"Our model line: {model_line_txt}")
            print(f"Edge: {abs(sig_spread['edge_pts']):.1f} points toward {rec_team}")
            print(f"Chance {rec_team} covers: {p_rec_covers*100:.1f}%")
            print(f"Break-even at {sig_spread['odds']}: {sig_spread['break_even']*100:.1f}%")
            print(f"Expected value: {sig_spread['ev_pct']*100:.1f}%")
            print(f"Recommendation: {sig_spread['signal']} → Bet {rec_team}")
            pred.update({
                "sig_spread_line": sig_spread["market_line"],
                "sig_spread_model": sig_spread["model_spread"],
                "sig_spread_edge_pts": sig_spread["edge_pts"],
                "sig_spread_p_home_covers": sig_spread["p_home_covers"],
                "sig_spread_ev_pct": sig_spread["ev_pct"],
                "sig_spread_side": rec_team,
                "sig_spread_reco": sig_spread["signal"],
            })

        if vegas_total is not None:
            sig_total = total_signal(pred["model_total_no_market"], vegas_total, DEFAULT_BOOK_ODDS)
            edge_dir = "Over" if sig_total["side"] == "OVER" else "Under"
            p_rec_hits = sig_total["p_over"] if sig_total["side"] == "OVER" else (1.0 - sig_total["p_over"])
            print("\n--- Total Signal ---")
            print(f"Vegas total: {vegas_total:.1f}")
            print(f"Our model total: {sig_total['model_total']:.1f}")
            print(f"Edge: {abs(sig_total['edge_pts']):.1f} points toward the {edge_dir}")
            print(f"Chance {edge_dir} hits: {p_rec_hits*100:.1f}%")
            print(f"Break-even at {sig_total['odds']}: {sig_total['break_even']*100:.1f}%")
            print(f"Expected value: {sig_total['ev_pct']*100:.1f}%")
            print(f"Recommendation: {sig_total['signal']} → Bet {edge_dir.upper()}")
            pred.update({
                "sig_total_line": sig_total["market_total"],
                "sig_total_model": sig_total["model_total"],
                "sig_total_edge_pts": sig_total["edge_pts"],
                "sig_total_p_over": sig_total["p_over"],
                "sig_total_ev_pct": sig_total["ev_pct"],
                "sig_total_side": edge_dir.upper(),
                "sig_total_reco": sig_total["signal"],
            })

        if odds_home is not None or odds_away is not None:
            sig_ml = moneyline_signal(sim["home_win_prob"], odds_home, odds_away)
            print("\n--- Moneyline Signals ---")
            for side, data in sig_ml.items():
                team = home.name if side == "HOME" else away.name
                print(f"{team}:")
                print(f"  Model win prob: {data['model_p']*100:.1f}%")
                print(f"  Odds: {data['odds']}")
                print(f"  Break-even: {data['break_even']*100:.1f}%")
                print(f"  Expected value: {data['ev_pct']*100:.1f}%")
                print(f"  Recommendation: {data['signal']}")
            pred.update({
                "sig_ml_home_model_p": sig_ml.get("HOME", {}).get("model_p"),
                "sig_ml_home_break_even": sig_ml.get("HOME", {}).get("break_even"),
                "sig_ml_home_ev_pct": sig_ml.get("HOME", {}).get("ev_pct"),
                "sig_ml_home_signal": sig_ml.get("HOME", {}).get("signal"),
                "sig_ml_away_model_p": sig_ml.get("AWAY", {}).get("model_p"),
                "sig_ml_away_break_even": sig_ml.get("AWAY", {}).get("break_even"),
                "sig_ml_away_ev_pct": sig_ml.get("AWAY", {}).get("ev_pct"),
                "sig_ml_away_signal": sig_ml.get("AWAY", {}).get("signal"),
            })

        if _ask_bool("\nSave this prediction to CSV?", True):
            maybe_save_outputs(pred)
    
    except Exception as e:
        print(f"\nError: {e}. Please try again.")
    input("Press Enter to exit...")  # Prevent IDE auto-closing

def main():
    print("College Football Score Predictor (interactive, situational + weather)")
    try:
        while True:
            run_once()
            if not _ask_bool("\nRun another matchup?", True):
                print("Done. Good luck!")
                break
    except KeyboardInterrupt:
        print("\nExiting gracefully. Good luck!")
    except Exception as e:
        print(f"\nError: {e}. Please try again.")
    input("Press Enter to exit...")  # Prevent IDE auto-closing

if __name__ == "__main__":
    main()