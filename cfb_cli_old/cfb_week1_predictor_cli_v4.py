# cfb_week1_predictor_cli_v2.py
# Week-1-ready CFB score predictor with interactive prompts, situational stats, and weather.

from dataclasses import dataclass
from typing import Optional, Dict, Tuple
import math
import csv
import os
import re
from datetime import datetime


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = load_workbook = Font = None  # Excel optional

# ---------- CONFIG: Global knobs (updated with 2023-2024 FBS data) ----------
# League baselines (sourced: TeamRankings, NCAA, SP+ 2023-2024)
AVG_OFF_EPA = 0.00;     STD_OFF_EPA = 0.14       # EPA/play, top ~+0.3, bottom ~-0.3
AVG_DEF_EPA = 0.00;     STD_DEF_EPA = 0.14
AVG_RET_PROD = 0.68;    STD_RET_PROD = 0.11      # SP+ returning production
AVG_COMPOSITE_100 = 50.0; STD_COMPOSITE_100 = 18.0  # 247Sports-style grades
AVG_QB_100 = 50.0;      STD_QB_100 = 20.0
AVG_OL_RET = 2.8;       STD_OL_RET = 1.2         # Returning OL starters
AVG_TEMPO = 65.8;       STD_TEMPO = 7.0          # Plays/game (NCAA 2023)

# Situational baselines
AVG_3D = 0.398;         STD_3D = 0.06            # 3rd-down conversion rate
AVG_RZ_TD = 0.615;      STD_RZ_TD = 0.075        # Red-zone TD rate
AVG_RZ_FG = 0.23;       STD_RZ_FG = 0.055        # Red-zone FG rate
AVG_FD = 20.2;          STD_FD = 3.5             # First downs per game
AVG_GIVE = 1.45;        STD_GIVE = 0.45          # Giveaways per game
AVG_TAKE = 1.45;        STD_TAKE = 0.45          # Takeaways per game

# Base scoring/pace
BASE_POINTS_PER_DRIVE = 2.28  # 2023 avg, down from 2.35 pre-2020
PLAYS_PER_DRIVE = 5.8
EFF_TO_PPD_K = 0.11  # Calibrated for ~5-7 pt swing per STD

# --- Betting model knobs ---
SPREAD_SIGMA_PTS = 13.0   # stdev of ATS error in points (CFB typical ~ 12-14)
TOTALS_SIGMA_PTS = 11.0   # stdev of total error (CFB typical ~ 9-12)
EV_THRESHOLDS = {         # EV% cutoffs for flags
    "strong": 0.05,       # >= +5% EV → STRONG
    "lean":   0.02        #  +2–5% EV → LEAN ; else PASS
}
DEFAULT_BOOK_ODDS = -110  # typical spread/total juice if you don't enter odds


# Weather sensitivities (adjusted for observed impacts)
WX_WIND_PPD_PER_MPH = -0.008   # ~8% PPD drop at 10mph
WX_WIND_TEMPO_PER_MPH = -0.0015
WX_RAIN_PPD = -0.06            # 6% PPD drop
WX_SNOW_PPD = -0.10            # 10% PPD drop
WX_RAIN_TEMPO = -0.035
WX_SNOW_TEMPO = -0.05
WX_COLD_PPD = -0.03            # Below 40°F
WX_HOT_PPD = -0.02             # Above 90°F

# Turnover margin nudge (pts per expected net TO)
MARGIN_PER_NET_TO = 3.2  # Calibrated from EPA studies (3-5 range)

# Home-field advantage
DEFAULT_HFA_POINTS = 2.2  # 2023 post-COVID avg

# Market anchoring
MARKET_BLEND = {"use": True, "spread_weight": 0.25, "total_weight": 0.25}

# Off/Def preseason weights
OFF_WEIGHTS = {"eff": 0.45, "ret": 0.15, "transfer": 0.08, "recruit": 0.05, "qb": 0.20, "ol": 0.05, "oc_cont": 0.05}
DEF_WEIGHTS = {"eff": 0.50, "ret": 0.20, "transfer": 0.08, "recruit": 0.05, "dc_cont": 0.10, "front7": 0.10}

@dataclass
class TeamInput:
    """Represents team data for prediction, including core and situational stats."""
    name: str
    off_epa: float = 0.0
    def_epa: float = 0.0
    returning_prod_off: float = AVG_RET_PROD
    returning_prod_def: float = AVG_RET_PROD
    transfer_grade_100: float = AVG_COMPOSITE_100
    recruit_grade_100: float = AVG_COMPOSITE_100
    qb_grade_100: float = AVG_QB_100
    ol_returning_starters: int = 3
    oc_continuity: bool = True
    dc_continuity: bool = True
    front7_continuity_100: Optional[float] = None
    tempo_plays_per_game: float = AVG_TEMPO
    third_down_off: float = AVG_3D
    third_down_def: float = AVG_3D
    rz_td_off: float = AVG_RZ_TD
    rz_fg_off: float = AVG_RZ_FG
    rz_td_def: float = AVG_RZ_TD
    rz_fg_def: float = AVG_RZ_FG
    first_downs_pg: float = AVG_FD
    giveaways_pg: float = AVG_GIVE
    takeaways_pg: float = AVG_TAKE

@dataclass
class GameWeather:
    """Represents weather conditions for a game."""
    is_indoor: bool = False
    temp_f: float = 70.0
    wind_mph: float = 5.0
    precip: str = "none"  # "none", "light", "rain", "snow"

# ---------- Core math ----------
def z(v: float, m: float, s: float) -> float:
    """Compute z-score for value v with mean m and std dev s."""
    s = s if s and s > 1e-9 else 1.0
    return (v - m) / s

def preseason_off_index(t: TeamInput) -> float:
    """Calculate offensive index using preseason components."""
    eff_z = z(t.off_epa, AVG_OFF_EPA, STD_OFF_EPA)
    ret_z = z(t.returning_prod_off, AVG_RET_PROD, STD_RET_PROD)
    tr_z = z(t.transfer_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    rec_z = z(t.recruit_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    qb_z = z(t.qb_grade_100, AVG_QB_100, STD_QB_100)
    ol_z = z(float(t.ol_returning_starters), AVG_OL_RET, STD_OL_RET)
    oc = 1.0 if t.oc_continuity else 0.0
    return (OFF_WEIGHTS["eff"]*eff_z + OFF_WEIGHTS["ret"]*ret_z + OFF_WEIGHTS["transfer"]*tr_z +
            OFF_WEIGHTS["recruit"]*rec_z + OFF_WEIGHTS["qb"]*qb_z + OFF_WEIGHTS["ol"]*ol_z +
            OFF_WEIGHTS["oc_cont"]*oc)

def preseason_def_index(t: TeamInput) -> float:
    """Calculate defensive index using preseason components."""
    eff_component = -t.def_epa
    eff_z = z(eff_component, -AVG_DEF_EPA, STD_DEF_EPA)
    ret_z = z(t.returning_prod_def, AVG_RET_PROD, STD_RET_PROD)
    tr_z = z(t.transfer_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    rec_z = z(t.recruit_grade_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    dc = 1.0 if t.dc_continuity else 0.0
    f7_z = 0.0 if t.front7_continuity_100 is None else z(t.front7_continuity_100, AVG_COMPOSITE_100, STD_COMPOSITE_100)
    return (DEF_WEIGHTS["eff"]*eff_z + DEF_WEIGHTS["ret"]*ret_z + DEF_WEIGHTS["transfer"]*tr_z +
            DEF_WEIGHTS["recruit"]*rec_z + DEF_WEIGHTS["dc_cont"]*dc + DEF_WEIGHTS["front7"]*f7_z)

def combined_plays(base_tempo_home: float, base_tempo_away: float, home_fd: float, away_fd: float, wx: GameWeather) -> float:
    """Estimate total plays based on team tempos, first downs, and weather."""
    raw = 0.96 * (base_tempo_home + base_tempo_away)
    fd_bump = 1.0 + 0.004 * ((home_fd - AVG_FD) + (away_fd - AVG_FD))
    raw *= max(0.85, min(1.15, fd_bump))
    if not wx.is_indoor:
        wind_factor = 1.0 + WX_WIND_TEMPO_PER_MPH * max(0.0, wx.wind_mph)
        raw *= max(0.85, wind_factor)
        if wx.precip in ("rain", "snow"):
            raw *= (1.0 + (WX_RAIN_TEMPO if wx.precip == "rain" else WX_SNOW_TEMPO))
    return raw

def weather_ppd_multiplier(wx: GameWeather) -> float:
    """Calculate PPD multiplier due to weather effects."""
    if wx.is_indoor:
        return 1.0
    mult = 1.0
    wind_pen = WX_WIND_PPD_PER_MPH * max(0.0, wx.wind_mph)
    mult *= max(0.70, 1.0 + wind_pen)
    if wx.precip == "rain":
        mult *= (1.0 + WX_RAIN_PPD)
    elif wx.precip == "snow":
        mult *= (1.0 + WX_SNOW_PPD)
    if wx.temp_f < 40:
        mult *= (1.0 + WX_COLD_PPD)
    elif wx.temp_f > 90:
        mult *= (1.0 + WX_HOT_PPD)
    return mult

def situational_ppd_multiplier(off: TeamInput, opp_def: TeamInput) -> float:
    """Calculate PPD multiplier from situational stats (3rd down, red zone, turnovers)."""
    def z(v, m, s): s = s if s and s > 1e-9 else 1.0; return (v - m) / s

    # 3rd down
    m_3d = 1.0 + 0.05 * (z(off.third_down_off, AVG_3D, STD_3D) - z(opp_def.third_down_def, AVG_3D, STD_3D))
    m_3d = max(0.85, min(1.15, m_3d))

    # Red zone (include non-scoring drives)
    e_off = 7.0 * max(0, min(1, off.rz_td_off)) + 3.0 * max(0, min(1, off.rz_fg_off))
    e_def_allowed = 7.0 * max(0, min(1, opp_def.rz_td_def)) + 3.0 * max(0, min(1, opp_def.rz_fg_def))
    rz_score_rate = (off.rz_td_off + off.rz_fg_off) / max(1e-6, opp_def.rz_td_def + opp_def.rz_fg_def)
    rz_factor = (e_off / 4.8) * rz_score_rate  # 4.8 ~ avg EP per RZ trip
    m_rz = max(0.85, min(1.15, rz_factor))

    # Turnover risk
    give_pen = max(0.0, z(off.giveaways_pg, AVG_GIVE, STD_GIVE))
    take_pen = max(0.0, z(opp_def.takeaways_pg, AVG_TAKE, STD_TAKE))
    m_to = max(0.90, 1.0 - 0.02 * give_pen - 0.015 * take_pen)

    return m_3d * m_rz * m_to

def expected_net_turnovers(home: TeamInput, away: TeamInput) -> float:
    """Estimate expected net turnovers (home advantage)."""
    home_net = (home.takeaways_pg - AVG_TAKE) - (home.giveaways_pg - AVG_GIVE)
    away_net = (away.giveaways_pg - AVG_GIVE) - (away.takeaways_pg - AVG_TAKE)
    return home_net + away_net

def model_points_per_drive(off_idx: float, opp_def_idx: float) -> float:
    """Calculate expected points per drive based on efficiency gap."""
    gap = off_idx - opp_def_idx
    return BASE_POINTS_PER_DRIVE * math.exp(EFF_TO_PPD_K * gap)

def predict_game(
    home: TeamInput,
    away: TeamInput,
    weather: GameWeather,
    use_preseason: bool = True,
    hfa_points: float = DEFAULT_HFA_POINTS,
    vegas_spread: Optional[float] = None,
    vegas_total: Optional[float] = None,
    use_market_blend: bool = MARKET_BLEND["use"],
    spread_weight: float = MARKET_BLEND["spread_weight"],
    total_weight: float = MARKET_BLEND["total_weight"],
) -> Dict[str, float]:
    """Predict game outcomes with scores, spread, and diagnostics."""
    # Off/Def indices
    if use_preseason:
        home_off_idx = preseason_off_index(home)
        home_def_idx = preseason_def_index(home)
        away_off_idx = preseason_off_index(away)
        away_def_idx = preseason_def_index(away)
    else:
        home_off_idx = z(home.off_epa, AVG_OFF_EPA, STD_OFF_EPA)
        away_off_idx = z(away.off_epa, AVG_OFF_EPA, STD_OFF_EPA)
        home_def_idx = z(-home.def_epa, -AVG_DEF_EPA, STD_DEF_EPA)
        away_def_idx = z(-away.def_epa, -AVG_DEF_EPA, STD_DEF_EPA)

    # Pace & drives
    min_tempo = 50.0  # Prevent div-by-zero
    plays = combined_plays(max(min_tempo, home.tempo_plays_per_game), max(min_tempo, away.tempo_plays_per_game),
                           home.first_downs_pg, away.first_downs_pg, weather)
    total_drives = max(10.0, plays / PLAYS_PER_DRIVE)
    den = max(1e-6, home.tempo_plays_per_game + away.tempo_plays_per_game)
    drives_home = total_drives * (home.tempo_plays_per_game / den)
    drives_away = total_drives - drives_home

    # Base PPD
    home_ppd = model_points_per_drive(home_off_idx, away_def_idx)
    away_ppd = model_points_per_drive(away_off_idx, home_def_idx)

    # Situational and weather multipliers
    wx_mult = weather_ppd_multiplier(weather)
    home_mult = situational_ppd_multiplier(home, away) * wx_mult
    away_mult = situational_ppd_multiplier(away, home) * wx_mult
    home_ppd *= home_mult
    away_ppd *= away_mult

    home_pts = home_ppd * drives_home
    away_pts = away_ppd * drives_away

    # Baseline model spread/total
    model_total = home_pts + away_pts
    model_spread = (home_pts - away_pts) + hfa_points

    # Turnover-margin nudge
    net_to = expected_net_turnovers(home, away)
    model_spread += MARGIN_PER_NET_TO * net_to

    # Re-split points
    home_pts_adj = (model_total + model_spread) / 2.0
    away_pts_adj = model_total - home_pts_adj

    final_spread = model_spread
    final_total = model_total
    if use_market_blend and (vegas_spread is not None or vegas_total is not None):
        if vegas_spread is not None:
            final_spread = (1 - spread_weight) * model_spread + spread_weight * vegas_spread
        if vegas_total is not None:
            final_total = (1 - total_weight) * model_total + total_weight * vegas_total
        home_pts_adj = (final_total + final_spread) / 2.0
        away_pts_adj = final_total - home_pts_adj

    home_pts_adj = max(0.0, home_pts_adj)
    away_pts_adj = max(0.0, away_pts_adj)

    return {
        "home_team": home.name,
        "away_team": away.name,
        "home_points": round(home_pts_adj, 1),
        "away_points": round(away_pts_adj, 1),
        "spread_home_minus_away": round(final_spread, 1),
        "total_points": round(final_total, 1),
        "model_spread_no_market": round(model_spread, 1),
        "model_total_no_market": round(model_total, 1),
        "vegas_spread": vegas_spread if vegas_spread is not None else "",
        "vegas_total": vegas_total if vegas_total is not None else "",
        "hfa_points": hfa_points,
        "used_preseason": use_preseason,
        "combined_plays_est": round(plays, 1),
        "home_ppd": round(home_ppd, 3),
        "away_ppd": round(away_ppd, 3),
        "drives_home": round(drives_home, 1),
        "drives_away": round(drives_away, 1),
        "net_turnovers_home_adv": round(net_to, 2),
        "wx_indoor": weather.is_indoor,
        "wx_temp_f": weather.temp_f,
        "wx_wind_mph": weather.wind_mph,
        "wx_precip": weather.precip,
    }

# ---------- CLI helpers ----------
def _ask(prompt: str, cast, default):
    """Ask for input, cast to type, enforce ranges, retry on bad input."""
    while True:
        raw = input(f"{prompt} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            val = cast(raw)
            if cast == float:
                if prompt.startswith("Returning production"):
                    return max(0.0, min(1.0, val))
                if prompt.startswith(("Transfer", "Recruiting", "QB", "Front-7")):
                    return max(0.0, min(100.0, val))
                if prompt.startswith("Offensive tempo"):
                    return max(50.0, val)
            elif cast == int and prompt.startswith("Returning OL"):
                return max(0, min(5, val))
            return val
        except ValueError:
            print("  ✖ Invalid input. Try again.")

import re

def _ask_spread(home_team_name: str, away_team_name: str):
    """
    Parse a point spread and return it as HOME minus AWAY.
    Accepted inputs:
      • 'Home -7.5' / '{home_team} -7.5'  → home favored by 7.5 → +7.5
      • 'Home +3'                         → home +3 dog         → -3.0
      • 'Away -3' / '{away_team} -3'     → away favored by 3   → -3.0
      • 'Away +6'                        → away +6 dog         → +6.0
      • Raw '+7.5' / '-3.5' (already Home–Away)
      • 'pk', 'pick', 'pickem', '0'      → 0.0
      • Enter to skip returns None
    Prints both the interpreted numeric value and a plain-English explanation.
    """

    def explain(val: float):
        if abs(val) < 1e-9:
            print("  → Interpreted as Home–Away = +0 (Pick’em)")
            print("    Explanation: Pick’em")
        elif val > 0:
            print(f"  → Interpreted as Home–Away = +{val:g}")
            print(f"    Explanation: {home_team_name} favored by {val:g}")
        else:
            print(f"  → Interpreted as Home–Away = {val:g}")
            print(f"    Explanation: {away_team_name} favored by {abs(val):g}")

    def has_word(text: str, word: str) -> bool:
        # Word-boundary match for single words; simple substring for multi-word team names
        word = word.strip().lower()
        text = text.lower()
        if " " in word:
            return word in text
        return re.search(r'\b' + re.escape(word) + r'\b', text) is not None

    while True:
        raw = input(
            f"Point spread (e.g., '{home_team_name} -7.5', '{away_team_name} +3', 'pk', or raw +/- like +7.5) [Enter to skip]: "
        ).strip()
        if raw == "":
            return None

        rl = raw.lower()

        # Pick’em shortcuts
        if rl in {"pk", "pick", "pickem", "pick'em", "0", "0.0"}:
            val = 0.0
            explain(val)
            return val

        # Extract a signed number
        m = re.search(r'([+-]?\d+(?:\.\d+)?)', rl)
        if not m:
            print("  ✖ Couldn't find a number. Try again.")
            continue
        num = float(m.group(1))

        # Detect which side was referenced (works with multi-word team names too)
        is_home = has_word(rl, "home") or has_word(rl, home_team_name)
        is_away = has_word(rl, "away") or has_word(rl, away_team_name)

        if is_home and is_away:
            print("  ✖ Please reference only one side (Home or Away or one team name), not both.")
            continue

        # No side tokens → treat number as already Home–Away
        if not (is_home or is_away):
            val = num
            explain(val)
            return val

        if is_home:
            # Home -7.5 → +7.5 ; Home +3 → -3
            val = +abs(num) if num < 0 else -abs(num)
            explain(val)
            return val

        if is_away:
            # Away -3 → -3 ; Away +6 → +6
            val = -abs(num) if num < 0 else +abs(num)
            explain(val)
            return val

def _ask_total():
    """
    Parse a game total. Accepts:
      • 'o/u 54.5', 'total 54.5', 'over 54.5', 'under 54.5'
      • '54.5'
      • Enter to skip (returns None)
    Retries until valid or Enter, and echoes the result.
    """
    while True:
        raw = input("Game total (e.g., 'o/u 54.5' or '54.5') [Enter to skip]: ").strip().lower()
        if raw == "":
            return None
        m = re.search(r'(\d+(?:\.\d+)?)', raw)
        if not m:
            print("  ✖ Couldn't find a number. Try again or press Enter to skip.")
            continue
        total = float(m.group(1))
        if total <= 0:
            print("  ✖ Total must be positive.")
            continue
        print(f"  → Interpreted total = {total:g}")
        return total


def _ask_bool(prompt: str, default: bool) -> bool:
    """Ask for Y/N input, return boolean."""
    dstr = "Y" if default else "N"
    while True:
        raw = input(f"{prompt} (Y/N) [{dstr}]: ").strip().lower()
        if raw == "":
            return default
        if raw in ("y", "yes"): return True
        if raw in ("n", "no"): return False
        print("  ✖ Please enter Y or N.")

def _ask_optional_float(prompt: str) -> Optional[float]:
    """Ask for an optional float, Enter = None, enforce reasonable ranges."""
    while True:
        raw = input(f"{prompt} (or Enter to skip): ").strip()
        if raw == "":
            return None
        try:
            val = float(raw)
            if "Vegas spread" in prompt:
                return max(-50.0, min(50.0, val))
            if "Vegas total" in prompt:
                return max(0.0, min(120.0, val))
            return val
        except ValueError:
            print("  ✖ Invalid number. Try again or press Enter to skip.")

def _rate(prompt: str, default: float) -> float:
    """Ask for a rate (0-1 or %), clamp to 0-1."""
    while True:
        raw = input(f"{prompt} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            v = float(raw)
            if v > 1.001:
                print(f"  ⚠ Assuming {v}% as {v/100}")
                v /= 100.0
            return max(0.0, min(1.0, v))
        except ValueError:
            print("  ✖ Invalid rate. Try again.")

def _ask_team(which: str) -> TeamInput:
    """Prompt for team data, including situational stats."""
    print(f"\n--- Enter {which} TEAM info ---")
    name = input("Team name: ").strip() or which
    off_epa = _ask("Last season OFF EPA/play (can be 0)", float, 0.0)
    def_epa = _ask("Last season DEF EPA/play (negative is better; can be 0)", float, 0.0)
    returning_prod_off = _ask("Returning production OFF (0-1)", float, AVG_RET_PROD)
    returning_prod_def = _ask("Returning production DEF (0-1)", float, AVG_RET_PROD)
    transfer_grade_100 = _ask("Transfer/portal impact grade (0-100)", float, AVG_COMPOSITE_100)
    recruit_grade_100 = _ask("Recruiting/roster quality grade (0-100)", float, AVG_COMPOSITE_100)
    qb_grade_100 = _ask("QB grade (0-100)", float, AVG_QB_100)
    ol_returning_starters = _ask("Returning OL starters (0-5)", int, 3)
    oc_continuity = _ask_bool("OC continuity from last year?", True)
    dc_continuity = _ask_bool("DC continuity from last year?", True)
    front7 = _ask_optional_float("Front-7 continuity grade (0-100)")
    tempo = _ask("Offensive tempo (plays/game, ~60-75)", float, AVG_TEMPO)
    third_off = _rate("Offense 3rd-down rate (0-1 or %)", AVG_3D)
    third_def = _rate("Defense 3rd-down allowed (0-1 or %)", AVG_3D)
    rz_td_off = _rate("Offense Red-Zone TD rate (0-1 or %)", AVG_RZ_TD)
    rz_fg_off = _rate("Offense Red-Zone FG rate (0-1 or %)", AVG_RZ_FG)
    rz_td_def = _rate("Defense RZ TD allowed (0-1 or %)", AVG_RZ_TD)
    rz_fg_def = _rate("Defense RZ FG allowed (0-1 or %)", AVG_RZ_FG)
    first_downs = _ask("First downs per game", float, AVG_FD)
    giveaways = _ask("Giveaways per game", float, AVG_GIVE)
    takeaways = _ask("Takeaways per game", float, AVG_TAKE)

    return TeamInput(
        name=name, off_epa=off_epa, def_epa=def_epa,
        returning_prod_off=returning_prod_off, returning_prod_def=returning_prod_def,
        transfer_grade_100=transfer_grade_100, recruit_grade_100=recruit_grade_100,
        qb_grade_100=qb_grade_100, ol_returning_starters=ol_returning_starters,
        oc_continuity=oc_continuity, dc_continuity=dc_continuity,
        front7_continuity_100=front7, tempo_plays_per_game=tempo,
        third_down_off=third_off, third_down_def=third_def,
        rz_td_off=rz_td_off, rz_fg_off=rz_fg_off, rz_td_def=rz_td_def, rz_fg_def=rz_fg_def,
        first_downs_pg=first_downs, giveaways_pg=giveaways, takeaways_pg=takeaways
    )

def _print_team_summary(t: TeamInput, which: str):
    """Print a summary of team inputs for review."""
    print(f"\n[{which} SUMMARY] {t.name}")
    print(f"  EPA: OFF {t.off_epa} | DEF {t.def_epa}")
    print(f"  Returning: OFF {t.returning_prod_off} | DEF {t.returning_prod_def}")
    print(f"  Transfer {t.transfer_grade_100} | Recruit {t.recruit_grade_100} | QB {t.qb_grade_100} | OL {t.ol_returning_starters}")
    print(f"  OC cont {t.oc_continuity} | DC cont {t.dc_continuity} | Front7 {t.front7_continuity_100}")
    print(f"  Tempo {t.tempo_plays_per_game}")
    print(f"  3rd: OFF {t.third_down_off} | DEF {t.third_down_def}")
    print(f"  RZ OFF: TD {t.rz_td_off} FG {t.rz_fg_off} | RZ DEF: TD {t.rz_td_def} FG {t.rz_fg_def}")
    print(f"  First downs {t.first_downs_pg} | Giveaways {t.giveaways_pg} | Takeaways {t.takeaways_pg}")

def _print_wx_summary(wx: GameWeather):
    """Print a summary of weather inputs for review."""
    print("\n[WEATHER SUMMARY]")
    print(f"  Indoor {wx.is_indoor} | Temp {wx.temp_f}F | Wind {wx.wind_mph} mph | Precip {wx.precip}")

def _ask_weather() -> GameWeather:
    """Prompt for weather data."""
    print("\n--- Weather / Venue ---")
    indoor = _ask_bool("Is the game indoors/under a roof?", False)
    temp = _ask("Temp (°F)", float, 70.0)
    wind = _ask("Wind (mph)", float, 5.0)
    precip = input("Precip (none/light/rain/snow) [none]: ").strip().lower() or "none"
    if precip not in ("none", "light", "rain", "snow"):
        precip = "none"
    return GameWeather(is_indoor=indoor, temp_f=temp, wind_mph=wind, precip=precip)

def maybe_save_outputs(pred: dict):
    """
    Save one prediction to a *daily* CSV (and pretty Excel if openpyxl is available).
    Filenames look like: cfb_predictions_YYYY-MM-DD.csv / .xlsx
    """
    # ---- Base folder (fixed path) ----
    folder = r"C:\Users\User\OneDrive\Desktop\CFP Betting Model\Scripts\PredictionResults"
    os.makedirs(folder, exist_ok=True)

    # ---- Daily filenames ----
    date_str = datetime.now().strftime("%Y-%m-%d")  # local date
    csv_path  = os.path.join(folder, f"cfb_predictions_{date_str}.csv")
    xlsx_path = os.path.join(folder, f"cfb_predictions_{date_str}.xlsx")

    # ---- Preferred front-of-sheet order (kept from before) ----
    preferred = [
        "away_team", "away_points",
        "home_team", "home_points",
        "spread_home_minus_away", "total_points",
        "vegas_spread", "vegas_total",
        "model_spread_no_market", "model_total_no_market",
        "hfa_points", "used_preseason",
        "net_turnovers_home_adv",
        "combined_plays_est", "home_ppd", "away_ppd",
        "drives_home", "drives_away",
        "wx_indoor", "wx_temp_f", "wx_wind_mph", "wx_precip",
    ]

    # ---- Build header dynamically: preferred keys first, then any remaining keys ----
    pred_keys = list(pred.keys())
    header = [k for k in preferred if k in pred_keys]
    header += [k for k in pred_keys if k not in header]

    # ---- Value formatter ----
    def fmt(x):
        if isinstance(x, float):
            return f"{x:.2f}"
        return x

    # ---- CSV append (create & write header if new) ----
    is_new_csv = not os.path.exists(csv_path)
    with open(csv_path, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        if is_new_csv:
            writer.writeheader()
        writer.writerow({k: fmt(pred.get(k, "")) for k in header})
    print(f"  ✓ Saved CSV to {csv_path}")

    # ---- Excel append (optional) ----
    if Workbook is None or load_workbook is None:
        return  # openpyxl not installed; CSV already saved

    # Create workbook if missing
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(xlsx_path)

    # If workbook exists but header changed, merge headers
    wb = load_workbook(xlsx_path)
    ws = wb.active
    existing_header = [c.value for c in ws[1]]
    if existing_header != header:
        union = existing_header[:] if existing_header else []
        for k in header:
            if k not in union:
                union.append(k)
        # Rebuild sheet with new header and re-write existing rows
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.remove(ws)
        ws = wb.create_sheet("Predictions", 0)
        ws.append(union)
        for r in rows:
            row_dict = dict(zip(existing_header, r))
            ws.append([fmt(row_dict.get(k, "")) for k in union])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        header = union

    # Append current row
    ws.append([fmt(pred.get(k, "")) for k in header])

    # Auto column widths
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    wb.save(xlsx_path)
    print(f"  ✓ Saved pretty Excel to {xlsx_path}")

import math

def _phi(x: float) -> float:
    "Standard normal CDF"
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))

def american_to_decimal(odds: float) -> float:
    if odds >= 100:
        return 1.0 + odds / 100.0
    elif odds <= -100:
        return 1.0 + 100.0 / abs(odds)
    else:
        raise ValueError("American odds should be ≤ -100 or ≥ +100")

def expected_value_pct(p: float, odds: float) -> float:
    """
    EV% on $1 stake at given American odds.
    EV = p*(decimal-1) - (1-p)*1
    """
    dec = american_to_decimal(odds)
    return p * (dec - 1.0) - (1.0 - p)

def tier_from_ev(ev_pct: float) -> str:
    if ev_pct >= EV_THRESHOLDS["strong"]:
        return "STRONG"
    if ev_pct >= EV_THRESHOLDS["lean"]:
        return "LEAN"
    return "PASS"

def spread_signal(model_spread: float, line_home_minus_away: float, odds: float = DEFAULT_BOOK_ODDS):
    """
    model_spread: your model's Home−Away spread (positive means home favored)
    line_home_minus_away: sportsbook line in Home−Away terms (what _ask_spread returns)
    """
    edge_pts = model_spread - line_home_minus_away
    z = edge_pts / SPREAD_SIGMA_PTS
    # Probability the HOME side covers the book line
    p_home_covers = _phi(z)

    # Compute EV for both sides at the given odds
    ev_home = expected_value_pct(p_home_covers, odds)
    ev_away = expected_value_pct(1.0 - p_home_covers, odds)

    # Choose the better side
    if ev_home > ev_away and ev_home > 0:
        side = "HOME"
        ev = ev_home
    elif ev_away > 0:
        side = "AWAY"
        ev = ev_away
    else:
        # both non-positive → PASS (default to the larger one for reporting)
        side = "HOME" if ev_home >= ev_away else "AWAY"
        ev = max(ev_home, ev_away)

    signal = tier_from_ev(abs(ev)) if ev > 0 else "PASS"
    return {
        "market_line": round(line_home_minus_away, 2),
        "model_spread": round(model_spread, 2),
        "edge_pts": round(edge_pts, 2),
        "p_home_covers": round(p_home_covers, 3),  # keep for prints (we’ll flip when recommending AWAY)
        "odds": odds,
        "ev_pct": round(ev, 3),
        "side": side,
        "signal": signal,
    }

def total_signal(model_total: float, line_total: float, odds: float = DEFAULT_BOOK_ODDS):
    edge_pts = model_total - line_total
    z = edge_pts / TOTALS_SIGMA_PTS
    # Probability the OVER hits vs the book total
    p_over = _phi(z)

    # EV for both directions
    ev_over = expected_value_pct(p_over, odds)
    ev_under = expected_value_pct(1.0 - p_over, odds)

    if ev_over > ev_under and ev_over > 0:
        side = "OVER"
        ev = ev_over
    elif ev_under > 0:
        side = "UNDER"
        ev = ev_under
    else:
        side = "OVER" if ev_over >= ev_under else "UNDER"
        ev = max(ev_over, ev_under)

    signal = tier_from_ev(abs(ev)) if ev > 0 else "PASS"
    return {
        "market_total": round(line_total, 1),
        "model_total": round(model_total, 1),
        "edge_pts": round(edge_pts, 1),
        "p_over": round(p_over, 3),  # keep for prints (we’ll flip when recommending UNDER)
        "odds": odds,
        "ev_pct": round(ev, 3),
        "side": side,
        "signal": signal,
    }

def moneyline_signal(model_home_win_prob: float, odds_home: float = None, odds_away: float = None):
    """
    Optional ML signaling if you compute model win prob.
    You can approximate from spread with a standard mapping later if you like.
    """
    out = {}
    if odds_home is not None:
        ev_h = expected_value_pct(model_home_win_prob, odds_home)
        out["HOME"] = {
            "model_p": round(model_home_win_prob, 3),
            "odds": odds_home,
            "ev_pct": round(ev_h, 3),
            "signal": tier_from_ev(ev_h) if ev_h > 0 else "PASS",
        }
    if odds_away is not None:
        p_away = 1.0 - model_home_win_prob
        ev_a = expected_value_pct(p_away, odds_away)
        out["AWAY"] = {
            "model_p": round(p_away, 3),
            "odds": odds_away,
            "ev_pct": round(ev_a, 3),
            "signal": tier_from_ev(ev_a) if ev_a > 0 else "PASS",
        }
    return out


def run_once():
    """Run a single prediction cycle with input, review, and output."""
    # ---- Collect inputs ----
    home = _ask_team("HOME")
    away = _ask_team("AWAY")
    wx = _ask_weather()

    # ---- Review & edit loop ----
    while True:
        print("\n=== Review your inputs ===")
        _print_team_summary(home, "HOME")
        _print_team_summary(away, "AWAY")
        _print_wx_summary(wx)
        choice = input("\nLooks good? Type 'ok' to continue, or type 'home', 'away', or 'wx' to re-enter that section (or 'restart' to start over): ").strip().lower()
        if choice in ("ok", ""):
            break
        elif choice == "home":
            home = _ask_team("HOME")
        elif choice == "away":
            away = _ask_team("AWAY")
        elif choice in ("wx", "weather"):
            wx = _ask_weather()
        elif choice == "restart":
            home = _ask_team("HOME")
            away = _ask_team("AWAY")
            wx = _ask_weather()
        else:
            print("  ✖ Not a valid option. Type ok/home/away/wx/restart.")

    # ---- Game context ----
    print("\n--- Game context ---")
    hfa = _ask("Home-field advantage in points", float, DEFAULT_HFA_POINTS)
    use_preseason = _ask_bool("Use preseason components (EPA/ret/recruit/etc.)?", True)
    use_market = _ask_bool("Blend with Vegas numbers if provided?", MARKET_BLEND["use"])

    vegas_spread = None
    vegas_total = None
    spread_w = MARKET_BLEND["spread_weight"]
    total_w  = MARKET_BLEND["total_weight"]

    if use_market:
        print("\nSpread entry is flexible. Examples:")
        print(f"  • '{home.name} -7.5'  (home favored by 7.5)")
        print(f"  • '{away.name} +3'    (away is +3 underdog → home favored by 3)")
        print("  • 'pk' or '0'         (pick'em)")
        print("  • Or raw Home–Away number: +7.5 or -3.5")
        vegas_spread = _ask_spread(home.name, away.name)
        vegas_total  = _ask_total()
        spread_w = _ask("Blend weight toward Vegas spread (0-1)", float, spread_w)
        total_w  = _ask("Blend weight toward Vegas total (0-1)", float, total_w)

    # ---- Run prediction ----
    pred = predict_game(
        home, away, wx,
        use_preseason=use_preseason,
        hfa_points=hfa,
        vegas_spread=vegas_spread, vegas_total=vegas_total,
        use_market_blend=use_market, spread_weight=spread_w, total_weight=total_w
    )

    # ---- Output: clearer sections ----
    print("\n=== Prediction ===")
    print(f"Final projected score: {pred['away_team']} {pred['away_points']} @ {pred['home_team']} {pred['home_points']}")

    print("\n--- Vegas-adjusted (if blended) ---")
    print(f"Spread (Home − Away): {pred['spread_home_minus_away']}   |   Total: {pred['total_points']}")

    print("\n--- Model only (no Vegas blend) ---")
    print(f"Spread (Home − Away): {pred['model_spread_no_market']}   |   Total: {pred['model_total_no_market']}")

    print("\n--- Diagnostics ---")
    print(f"Combined plays est: {pred['combined_plays_est']}")
    print(f"Home PPD: {pred['home_ppd']}   |   Away PPD: {pred['away_ppd']}")
    print(f"Drives: home {pred['drives_home']}  |  away {pred['drives_away']}")
    print(f"Expected net TO (home advantaged): {pred['net_turnovers_home_adv']}")
    print(f"WX: indoor={pred['wx_indoor']} temp={pred['wx_temp_f']}F wind={pred['wx_wind_mph']}mph precip={pred['wx_precip']}")

    # ---- Betting signals (friendly phrasing) ----
    try:
        if vegas_spread is not None:
            sig_spread = spread_signal(pred["model_spread_no_market"], vegas_spread, DEFAULT_BOOK_ODDS)

            vegas_fav = home.name if vegas_spread > 0 else away.name
            vegas_line_txt = f"{vegas_fav} {abs(vegas_spread):.1f}"

            model_fav = home.name if sig_spread["model_spread"] > 0 else away.name
            model_line_txt = f"{model_fav} {abs(sig_spread['model_spread']):.1f}"

            rec_team = home.name if sig_spread["side"] == "HOME" else away.name
            p_rec_covers = sig_spread["p_home_covers"] if sig_spread["side"] == "HOME" else (1.0 - sig_spread["p_home_covers"])

            print("\n--- Spread Signal ---")
            print(f"Vegas line: {vegas_line_txt}")
            print(f"Our model line: {model_line_txt}")
            print(f"Edge: {abs(sig_spread['edge_pts']):.1f} points toward {rec_team}")
            print(f"Chance {rec_team} covers: {p_rec_covers*100:.1f}%")
            print(f"Expected value: {sig_spread['ev_pct']*100:.1f}%")
            print(f"Recommendation: {sig_spread['signal']} → Bet {rec_team}")

            # Optional: attach to pred so it saves
            pred.update({
                "sig_spread_line": sig_spread["market_line"],
                "sig_spread_model": sig_spread["model_spread"],
                "sig_spread_edge_pts": sig_spread["edge_pts"],
                "sig_spread_p_home_covers": sig_spread["p_home_covers"],
                "sig_spread_ev_pct": sig_spread["ev_pct"],
                "sig_spread_side": rec_team,
                "sig_spread_reco": sig_spread["signal"],
            })

        if vegas_total is not None:
            sig_total = total_signal(pred["model_total_no_market"], vegas_total, DEFAULT_BOOK_ODDS)

            edge_dir = "Over" if sig_total["side"] == "OVER" else "Under"
            p_rec_hits = sig_total["p_over"] if sig_total["side"] == "OVER" else (1.0 - sig_total["p_over"])

            print("\n--- Total Signal ---")
            print(f"Vegas total: {vegas_total:.1f}")
            print(f"Our model total: {sig_total['model_total']:.1f}")
            print(f"Edge: {abs(sig_total['edge_pts']):.1f} points toward the {edge_dir}")
            print(f"Chance {edge_dir} hits: {p_rec_hits*100:.1f}%")
            print(f"Expected value: {sig_total['ev_pct']*100:.1f}%")
            print(f"Recommendation: {sig_total['signal']} → Bet {edge_dir.upper()}")

            # Optional: attach to pred so it saves
            pred.update({
                "sig_total_line": sig_total["market_total"],
                "sig_total_model": sig_total["model_total"],
                "sig_total_edge_pts": sig_total["edge_pts"],
                "sig_total_p_over": sig_total["p_over"],
                "sig_total_ev_pct": sig_total["ev_pct"],
                "sig_total_side": edge_dir.upper(),
                "sig_total_reco": sig_total["signal"],
            })
    except NameError:
        # If spread_signal / total_signal aren't defined, skip gracefully
        pass

    # ---- Save to file ----
    if _ask_bool("\nSave this prediction to CSV?", True):
        maybe_save_outputs(pred)


def main():
    """Main loop for running predictions."""
    print("College Football Score Predictor (interactive, situational + weather)")
    try:
        while True:
            run_once()
            if not _ask_bool("\nRun another matchup?", True):
                print("Done. Good luck!")
                break
    except KeyboardInterrupt:
        print("\nExiting gracefully. Good luck!")
    except Exception as e:
        print(f"\nError: {e}. Please try again.")

if __name__ == "__main__":
    main()